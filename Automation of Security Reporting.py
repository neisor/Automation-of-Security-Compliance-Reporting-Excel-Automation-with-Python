import openpyxl
import pyautogui
from openpyxl import Workbook
from openpyxl import load_workbook
import clipboard
import time
import tkinter
from tkinter.filedialog import askopenfilename
from tkinter import filedialog
import datetime

#FailSafe
pyautogui.FAILSAFE = True

#Define global filePath variable
global filePath
filePath = ''

global outputPath
outputPath = ''

### select EXCEL FUNCTION STARTS HERE
def selectExcel():
    try:
        global filePath
        filePath = askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *.xlsb")]) # show an "Open" dialog box and return the path to the selected file
        filePath = filePath.replace("/", "\\")
    
    except:
        pyautogui.alert(title='Error', text='Seems like no Excel sheet has been selected. Please, select one...')

### MAIN FUNCTION STARTS HERE
def mainFunction():
    global filePath
    print (filePath)

    global outputPath

    if (outputPath is '' or filePath is ''):
        pyautogui.alert(title='Error', text="You did not select the Excel document, or, you did not select the location for the outputted file. Select what's needed and try again.")
        return

    #List with all the CMO+ servers
    cmoPlusServers = ["server1", "server2", "server3"]
        
    try:
        #WORKING WITH EXCEL
        wb = load_workbook(filename=str(filePath), read_only=True)
        ws = wb['PC Details']
        row_count = ws.max_row
        column_count = ws.max_column
        print('Row count is: ' + str(row_count))
        print('Column count is: ' + str(column_count))
        
    except:
        pyautogui.alert(title='Error', text='Something went wrong during the initialization of the existing Excel file.')

    ### CREATE AND POPULATE THE WRITE_ONLY EXCEL SHEET
    try:
        wb_output = Workbook(write_only=True)
        ws_output = wb_output.create_sheet(title='PC Details')

        #Define the rowIdentificator variable
        rowIdentificator = 0

        print('ws.rows is: ' + str(ws.rows))
        
        #Populate the main PC Details sheet with servers and formulas
        for row in ws.rows:
            rowIdentificator += 1
            rowInCorrectFormat = [cell.value for cell in row]
            if rowIdentificator is 5:
               rowInCorrectFormat.append('OK or NOT COMPLIANT WITH OUR SECURITY STANDARDS?')
               rowInCorrectFormat.append('Is this a CMO+ server?')
            elif rowIdentificator > 5:
                 rowInCorrectFormat.append('=IF(H' + str(rowIdentificator) + '=I' + str(rowIdentificator) + ',"OK","NOT COMPLIANT WITH OUR SECURITY STANDARDS")')
                 rowInCorrectFormat.append("=VLOOKUP(B" + str(rowIdentificator) + ",'List of CMO+ servers'!A$1:B$98,2,FALSE)")

            ws_output.append(rowInCorrectFormat)
        
        #Create List of CMO+ servers sheet
        ws_output = wb_output.create_sheet(title='List of CMO+ servers')
         
    except:
        pyautogui.alert(title='Error', text='Something went wrong during the creation of a new Excel file.')

    #Populate the List of CMO+ servers sheet and SAVE THE WHOLE OUTPUTTED EXCEL
    try:
        for server in cmoPlusServers:
            rowInCorrectFormat = [server, 'CMO+ server']
            ws_output.append(rowInCorrectFormat)
  
        #Get the actual time and format it
        actualTime = time.time()
        convertedActualTime = datetime.datetime.fromtimestamp(actualTime).strftime('%d_%m_%Y-%H_%M_%S')

        savedPathOfOutputtedFile = (str(outputPath) + '\ComplianceReportAutomation_' + str(convertedActualTime) + '.xlsx')
        wb_output.save(str(savedPathOfOutputtedFile))
        
    except:
        pyautogui.alert(title='Error', text='Something went wrong during the population of the List of CMO+ Servers sheet in the created Excel file. Try again.')

    #Show a success message once completed
    pyautogui.alert(title='Success', text='Successfully finished.\nYou can find the outputted Excel file at:\n\n' + str(savedPathOfOutputtedFile))

#Define function for choosing the output location of the outputted Excel file
def selectOutputLocation():

    global outputPath

    #Ask for a directory where to put the outputted Excel file
    outputPath = filedialog.askdirectory()
    outputPath = outputPath.replace("/", "\\")
    
#GUI
top = tkinter.Tk()

top.title('ComplRepAuto')
top.geometry('340x240')

labelTitle = tkinter.Label(text='Compliance Report Automation\n')
labelTitle.grid(row = 1, column=1, columnspan=6)

button = tkinter.Button(text = 'Start!', command = mainFunction)
button.grid(row = 2, column=1, columnspan=6)

buttonSelect = tkinter.Button(text = 'Select Excel', command = selectExcel)
buttonSelect.grid(row = 3, column=1, columnspan=6)

buttonSelectOutput = tkinter.Button(text = 'Select Output Location', command = selectOutputLocation)
buttonSelectOutput.grid(row = 4, column=1, columnspan=6)

labelCreatedBy = tkinter.Label(text='\n                                Created by                                \nAntonio Raffaele Iannaccone\nantonio-raffaele.iannaccone@t-systems.com')
labelCreatedBy.grid(row = 5, column=1, columnspan=6)

top.mainloop()
